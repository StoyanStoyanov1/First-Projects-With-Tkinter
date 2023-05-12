from tkinter import *
from openpyxl import load_workbook
from tkinter import messagebox

import re


def letters_and_digits(text):
    pattern = r'^[a-zA-Z0-9]+$'
    match = re.match(pattern, text)
    if match:
        return True
    else:
        return False


def clear_entry(event):
    event.widget.delete(0, END)


def password_validator(received_password):
    result = []
    password_is_valid = True
    digit_counter = 0
    for character in received_password:
        if character.isdigit():
            digit_counter += 1

    if not 6 <= len(received_password) <= 10:
        result.append('Password must be between 6 and 10 characters')
        password_is_valid = False
    if not received_password.isalnum():
        result.append('Password must consist only of letters and digits')
        password_is_valid = False
    if not digit_counter >= 2:
        result.append('Password must have at least 2 digits')
        password_is_valid = False

    if password_is_valid:
        return 'Password is valid'
    else:
        return '\n'.join(result)


def add_entry(event):
    if len(username_entry.get()) == 0:
        username_entry.insert(0, "Your account")

    if len(password_entry.get()) == 0:
        password_entry.insert(0, "Your password")

    if len(repeat_password_entry.get()) == 0:
        repeat_password_entry.insert(0, "Repeat your password")


def new_account():
    name = username_entry.get()
    if name == "Your account":
        return messagebox.showerror("Invalid", "Add your account name")
    if not (len(name) > 2 and letters_and_digits(name)):
        return messagebox.showerror("Invalid", "Your username must be at least "
                                               "3 characters long and contain only letters and numbers")
    for row in registers.iter_rows():
        username, password = (cell.value for cell in row)
        if name == username:
            return messagebox.showerror("Invalid", "Тhis username already exists")
    password = password_entry.get()
    if password == "Your password":
        return messagebox.showerror("Invalid", "Add your password")
    current_password = password_validator(password)
    if current_password != "Password is valid":
        return messagebox.showerror("Invalid", f"{current_password}")
    repeat_password = repeat_password_entry.get()
    if repeat_password == "Repeat your password":
        return messagebox.showerror("Invalid", "Repeat your password")
    if repeat_password != password:
        return messagebox.showerror("Invalid", "The passwords do not match")
    else:
        registers.append([name, password])

        file.save("register.xlsx")
        return messagebox.showinfo("Valid", "You have a new account!")


file = load_workbook(filename="register.xlsx")
registers = file.active

new_window = Tk()
new_window.title("Register")
new_window.geometry("877x493")
new_window.iconbitmap("favicon.ico")
new_window.wm_maxsize(877, 493)
new_background_image = PhotoImage(file="login.png")
new_background_label = Label(new_window, image=new_background_image)
new_background_label.place(x=0, y=0, relheight=1, relwidth=1)

icon_account = PhotoImage(file="account.png")
icon_password = PhotoImage(file="password.png")

username_entry = Entry(new_window, font=("Arial", 14))
username_entry.place(x=340, y=120)
username_entry.insert(0, "Your account")
username_entry.bind("<FocusIn>", clear_entry)
login_icon_label = Label(new_window, image=icon_account, bg='white')
login_icon_label.place(x=310, y=120)
username_entry.bind("<FocusOut>", add_entry)

password_entry = Entry(new_window, font=("Arial", 14))
password_entry.place(x=340, y=180)
password_icon_label = Label(new_window, image=icon_password, bg="white")
password_icon_label.place(x=310, y=180)
password_entry.insert(0, "Your password")
password_entry.bind("<FocusIn>", clear_entry)
password_entry.bind("<FocusOut>", add_entry)

repeat_password_entry = Entry(new_window, font=("Arial", 14))
repeat_password_entry.place(x=340, y=240)
repeat_password_icon_label = Label(new_window, image=icon_password, bg="white")
repeat_password_icon_label.place(x=310, y=240)
repeat_password_entry.insert(0, "Repeat your password ")
repeat_password_entry.bind("<FocusIn>", clear_entry)
repeat_password_entry.bind("<FocusOut>", add_entry)

password_button = Button(new_window, text="Register", font=("bold", 13), width=15, command=new_account)
password_button.place(x=366, y=300)

new_window.mainloop()
