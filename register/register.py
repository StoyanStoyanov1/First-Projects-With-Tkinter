from tkinter import *
from openpyxl import load_workbook
from tkinter import messagebox


def new_register():
    import subprocess
    subprocess.Popen(['python', 'new_register.py'])


def clear_entry(event):
    event.widget.delete(0, END)


def add_entry(event):
    if len(username_entry.get()) == 0:
        username_entry.insert(0, "Your account")

    if len(password_entry.get()) == 0:
        password_entry.insert(0, "Your password")


def login():
    for row in registers.iter_rows():
        username, password = (cell.value for cell in row)
        if username_entry.get() == username:

            if password_entry.get() == password:
                messagebox.showinfo("Valid", "Wellcome!")
                break
            else:
                messagebox.showerror("Invalid", "Invalid password")
                break
    else:
        messagebox.showerror("Invalid", "Invalid username")

        username_entry.delete("0", END)
        password_entry.delete("0", END)


file = load_workbook(filename="register.xlsx")
registers = file.active

root = Tk()
root.title("Register")
root.geometry("877x493")
root.iconbitmap("favicon.ico")
root.wm_maxsize(877, 493)
background_image = PhotoImage(file="login.png")
background_label = Label(root, image=background_image)
background_label.place(x=0, y=0, relheight=1, relwidth=1)

icon_account = PhotoImage(file="account.png")
icon_password = PhotoImage(file="password.png")

username_entry = Entry(root, font=("Arial", 14))
username_entry.place(x=340, y=120)
username_entry.insert(0, "Your account")
username_entry.bind("<FocusIn>", clear_entry)
login_icon_label = Label(root, image=icon_account, bg='white')
login_icon_label.place(x=310, y=120)
username_entry.bind("<FocusOut>", add_entry)

password_entry = Entry(root, font=("Arial", 14))
password_entry.place(x=340, y=180)
password_icon_label = Label(root, image=icon_password, bg="white")
password_icon_label.place(x=310, y=180)
password_entry.insert(0, "Your password")
password_entry.bind("<FocusIn>", clear_entry)
password_entry.bind("<FocusOut>", add_entry)

login_button = Button(root, text="Login", font=("bold", 13), width=15, command=login)
login_button.place(x=366, y=240)
register_button = Button(root, text="Register", font=("bold", 13), width=15, command=new_register)
register_button.place(x=366, y=300)

root.mainloop()
